from flask import Flask, render_template, request, jsonify
from flask_session import Session
import os
import numpy as np
import logging
import pandas as pd
from datetime import datetime
from tensorflow.keras.models import load_model
import openpyxl

app = Flask(__name__, static_folder='static')
app.secret_key = 'gandalf'
app.config['SESSION_TYPE'] = 'filesystem'

# Load the past predictions from the Excel file
script_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(script_dir, 'predict.xlsx')

# Check if the predict.xlsx file exists
if not os.path.exists(excel_path):
    # Create an empty DataFrame with the desired columns
    initial_df = pd.DataFrame(columns=['Date', 'Num_Predictions', 'Success_Count'])

    # Save the DataFrame to the Excel file
    initial_df.to_excel(excel_path, index=False)
    print(f"Created {excel_path}")

    # Save the DataFrame to the Excel file
    initial_df.to_excel(excel_path, index=False)
    print(f"Created {excel_path}")

# Load the DataFrame from the Excel file
try:
    comparison_df = pd.read_excel(excel_path)
except Exception as e:
    logging.error(f"Error loading comparison DataFrame from Excel file: {e}")
    comparison_df = pd.DataFrame(columns=['Date', 'Selected_Numbers', '1_success', '2_success', '3_success', '4_success',
                                           '5_success', '6_success', '7_success', '8_success', '9_success', '10_success',
                                           '11_success', '12_success'])

# Define comparison_result_dict as an empty dictionary
comparison_result_dict = {}


def predict_numbers(input_numbers, num_predictions_list=[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]):
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        model_path = os.path.join(script_dir, 'prediction_model_lstm')
        model = load_model(model_path, compile=False)  # Load model without compiling

        logging.debug(f"Input numbers shape: {input_numbers.shape}")
        logging.debug(f"Input numbers values: {input_numbers}")

        prediction = model.predict(input_numbers)

        logging.debug(f"Raw prediction shape: {prediction.shape}")
        logging.debug(f"Raw prediction values: {prediction}")

        predictions_scalar = (prediction.flatten() * 79) + 1

        logging.debug(f"Predicted scalar values: {predictions_scalar}")

        clipped_predictions = np.clip(predictions_scalar, 1, 80)

        logging.debug(f"Clipped prediction values: {clipped_predictions}")

        rounded_predictions = np.floor(clipped_predictions).astype(int)

        logging.debug(f"Rounded prediction values: {rounded_predictions}")

        unique_predictions = list(set(rounded_predictions))

        logging.debug(f"Unique predictions before additional: {unique_predictions}")

        while len(unique_predictions) < max(num_predictions_list):
            additional_predictions = np.random.choice(
                np.setdiff1d(np.arange(1, 81), unique_predictions),
                max(num_predictions_list) - len(unique_predictions),
                replace=False
            )
            unique_predictions.extend(additional_predictions)

        logging.debug(f"Final unique predictions: {unique_predictions}")

        unique_predictions.sort()

        logging.debug(f"Final sorted predictions: {unique_predictions}")

        predictions_dict = {num_predictions: unique_predictions[:num_predictions] for num_predictions in num_predictions_list}

        return predictions_dict
    except Exception as e:
        logging.error(f"Error during prediction: {e}")
        return f"Error during prediction: {e}"

def update_excel_file(selected_numbers, success_counts):
    try:
        global comparison_df
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Create a new DataFrame with the current time and selected numbers
        new_row = {'Date': current_time, 'Selected_Numbers': str(selected_numbers)}

        # Add separate columns for success counts
        for num_predictions in range(1, 13):
            new_row[f'{num_predictions}_success'] = success_counts.get(num_predictions, 0)

        new_df = pd.DataFrame([new_row])

        # Append the new DataFrame to the existing one
        comparison_df = pd.concat([comparison_df, new_df], ignore_index=True)

        # Save the updated DataFrame to the Excel file
        comparison_df.to_excel(excel_path, index=False)

        logging.info(f"Updated Excel file with new results at {current_time}")
        print({'Date': current_time, 'Selected_Numbers': str(selected_numbers), 'Success_Counts': success_counts})
    except Exception as e:
        logging.error(f"Error updating Excel file: {e}")
        print(f"Error updating Excel file: {e}")

@app.route('/', methods=['GET', 'POST'])
def index():
    global comparison_result_dict, saved_forecast

    if request.method == 'POST':
        selected_numbers = request.form.getlist('selected_numbers[]')

        # Check if selected_numbers is not empty
        if not selected_numbers:
            return render_template('index.html', prediction='Error: No numbers selected.')

        # Convert selected numbers to integers
        selected_numbers = [int(number) for number in selected_numbers]

        # Convert the list of selected numbers to a NumPy array
        selected_numbers_array = np.array(selected_numbers)

        # Reshape the array for prediction
        selected_numbers_reshaped = selected_numbers_array.reshape(-1, 1)

        # Make predictions using the loaded model for different sizes
        predictions = predict_numbers(selected_numbers_reshaped, num_predictions_list=[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12])

        # Create a dictionary to hold your predictions
        current_forecast = {
            1: predictions[1],
            2: predictions[2],
            3: predictions[3],
            4: predictions[4],
            5: predictions[5],
            6: predictions[6],
            7: predictions[7],
            8: predictions[8],
            9: predictions[9],
            10: predictions[10],
            11: predictions[11],
            12: predictions[12],
        }

        # Compare with the saved forecast if available
        if 'saved_forecast' in globals() and saved_forecast:
            comparison_result_dict = compare_forecasts(saved_forecast, selected_numbers)
            update_excel_file(selected_numbers, calculate_success_counts(comparison_result_dict))

        # Update saved_forecast with the current predictions
        saved_forecast = current_forecast.copy()

        return render_template('index.html', prediction_dict=current_forecast, comparison_result_dict=comparison_result_dict, selected_numbers=selected_numbers)

    return render_template('index.html', prediction_dict={}, comparison_result_dict={}, selected_numbers=[])

def compare_forecasts(forecast, selected_numbers):
    comparison_result_dict = {}
    success_counts = {num_predictions: 0 for num_predictions in range(1, 13)}

    for num_predictions in range(1, 13):
        try:
            if forecast[num_predictions]:
                comparison_result_dict[num_predictions] = []
                for prediction in forecast[num_predictions]:
                    match = 1 if prediction in selected_numbers else 0
                    comparison_result_dict[num_predictions].append({'number': prediction, 'tie': match})
                    success_counts[num_predictions] += match

        except Exception as e:
            logging.error(f"Error during comparison for num_predictions={num_predictions}: {e}")

    return comparison_result_dict


def calculate_success_counts(comparison_result_dict):
    success_counts = {num_predictions: 0 for num_predictions in range(1, 13)}

    for num_predictions, predictions in comparison_result_dict.items():
        for prediction in predictions:
            success_counts[num_predictions] += prediction['tie']

    return success_counts

@app.route('/predictions', methods=['GET'])
def predictions():
    # Generate 10 sets of predictions for the whole day
    daily_predictions = {}
    for i in range(1, 11):
        # Generate random input sequence (you may adjust this based on your actual input)
        random_input_sequence = np.random.randint(1, 81, size=(20,))
        random_input_sequence_reshaped = random_input_sequence.reshape(-1, 1)

        # Make predictions using the loaded model
        predictions = predict_numbers(random_input_sequence_reshaped, num_predictions_list=[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12])

        # Select specific names from the predictions (modify this based on your requirements)
        selected_predictions = predictions[1]

        # Call save_predictions to save the selected predictions
        save_predictions(1, selected_predictions)  # Pass both parameters here

        daily_predictions[f"Prediction {i}"] = predictions

    return render_template('predictions.html', daily_predictions=daily_predictions)


def save_predictions(num_predictions, numbers):
    try:
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Open the Excel file
        wb = openpyxl.load_workbook(excel_path)

        # Select the active sheet
        sheet = wb.active

        # Find the first empty row in column A
        row_number = 1
        while sheet.cell(row=row_number, column=1).value:
            row_number += 1

        # Update the cells with the current time and predictions
        sheet.cell(row=row_number, column=1, value=current_time)
        sheet.cell(row=row_number, column=2, value=str(num_predictions))

        for index, prediction in enumerate(numbers, start=1):
            sheet.cell(row=row_number, column=index + 2, value=str(prediction))

        # Save the changes
        wb.save(excel_path1)

        logging.info(f"Updated Excel file with new results at {current_time}")
        print({'Date': current_time, 'num_predictions': str(num_predictions), 'numbers': numbers})
    except Exception as e:
        logging.error(f"Error updating Excel file: {e}")
        print(f"Error updating Excel file: {e}")

@app.route('/save_predictions', methods=['POST'])
def save_selected_predictions():
    try:
        data = request.get_json()

        # Extract selected predictions and sheet name from the request data
        selected_predictions = data.get('selected_predictions', [])
        sheet_name = data.get('sheet_name', '')

        # Ensure the variable is defined before using it
        selected_predictions_df = pd.DataFrame()

        # Iterate over selected predictions and create a DataFrame
        for index, selected_prediction in enumerate(selected_predictions, start=1):
            num_predictions = selected_prediction['num_predictions']
            numbers = selected_prediction['numbers']

            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Create a new row for each prediction set
            new_row = {'Date': current_time, 'num_predictions': str(num_predictions)}

            # Add separate columns for success counts
            for index, prediction in enumerate(numbers, start=1):
                new_row[f'{index}_numbers'] = prediction

            # Concatenate the new row to the DataFrame
            selected_predictions_df = pd.concat([selected_predictions_df, pd.DataFrame([new_row])], ignore_index=True)

        # Save the selected predictions DataFrame to the Excel file
        selected_predictions_df.to_excel(excel_path1, index=False)

        logging.info(f"Saved selected predictions to Excel file at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logging.info(f"Saved numbers list {num_predictions}")
        return jsonify({"message": "Selected predictions saved successfully!"}), 200

    except Exception as e:
        logging.error(f"Error saving selected predictions: {e}")
        return jsonify({"error": f"Error saving selected predictions: {e}"}), 500



if __name__ == '__main__':
    app.run(debug=True)
